from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NUMBER_FORMAT = "#,##0"


def _safe_title(name: str, used: set[str]) -> str:
    title = "".join(c for c in name if c not in "[]:*?/\\")[:31] or "Sheet"
    base, n = title, 2
    while title.lower() in used:
        suffix = f" {n}"
        title = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(title.lower())
    return title


def build_workbook(sheets: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()

    for sheet in sheets:
        columns = sheet.get("columns", [])
        rows = sheet.get("rows", [])
        ws = wb.create_sheet(_safe_title(sheet.get("name", "Sheet"), used))

        ws.append([c["label"] for c in columns])
        for row in rows:
            ws.append([row.get(c["key"]) for c in columns])

        for idx, col in enumerate(columns, start=1):
            letter = get_column_letter(idx)
            width = max(len(col["label"]) + 4, 12)
            ws.column_dimensions[letter].width = min(width, 40)
            if col.get("type") == "number":
                for r in range(2, ws.max_row + 1):
                    cell = ws[f"{letter}{r}"]
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = NUMBER_FORMAT

        for idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=idx)
            cell.fill, cell.font = HEADER_FILL, HEADER_FONT

        ws.freeze_panes = "A2"
        last_data_row = ws.max_row
        if last_data_row < 2:
            continue
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{last_data_row}"

        total_row = last_data_row + 1
        for idx, col in enumerate(columns, start=1):
            letter = get_column_letter(idx)
            if col.get("type") == "number":
                cell = ws.cell(row=total_row, column=idx,
                               value=f"=SUBTOTAL(109,{letter}2:{letter}{last_data_row})")
                cell.number_format = NUMBER_FORMAT
            elif idx == 1:
                ws.cell(row=total_row, column=idx, value="Total")
            cell = ws.cell(row=total_row, column=idx)
            cell.fill, cell.font = HEADER_FILL, HEADER_FONT

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
